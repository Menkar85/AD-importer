import openpyxl
from openpyxl.styles import Font, Alignment
import transliterate
import logging

logger = logging.getLogger(__name__)


def get_excel_data(path, *, has_headers=True) -> dict:
    """Extract data from an Excel workbook.

    Loads an Excel file and extracts headers and data rows. Supports both
    files with and without headers.

    Args:
        path (str): Path to the Excel file to load.
        has_headers (bool, optional): Whether the first row contains headers.
            Defaults to True.

    Returns:
        dict: Dictionary containing:
            - 'headers': List of header strings (empty if has_headers=False)
            - 'data': List of data rows, each row is a list of cell values

    Raises:
        Exception: If the workbook cannot be loaded.
        ValueError: If the Excel file contains no data.

    Note:
        If has_headers is True, the first row is treated as headers and data
        starts from row 2. If False, all rows are treated as data.
    """
    try:
        wb = openpyxl.load_workbook(path)
        logger.debug(f"Workbook {path} loaded.")
        sheet = wb.active
    except Exception as e:
        logger.critical(f"Unable to load workbook")
        raise e

    if has_headers and sheet.max_row > 0:
        headers = next(sheet.iter_rows(min_row=1, values_only=True))
    elif sheet.max_row > 0:
        headers = []
    else:
        raise ValueError("Excel book has no data!")

    start_row = 2 if has_headers else 1
    data_rows = list(sheet.iter_rows(min_row=start_row, values_only=True))
    return {"headers": headers, "data": data_rows}


def build_import_data(input_dict, upn, mail_domain=None) -> list[dict]:
    """Build user data dictionaries from Excel input data.

    Converts raw Excel data into structured dictionaries suitable for AD user creation.
    Transliterates Russian names to Latin characters and constructs user attributes.

    Args:
        input_dict (dict): Dictionary containing 'headers' and 'data' from Excel file.
        upn (str): User Principal Name suffix (e.g., "example.com").
        mail_domain (str, optional): Domain for email addresses. If None, uses upn.

    Returns:
        list[dict]: List of user dictionaries, each containing:
            - cn: Common name (username)
            - givenName: First name
            - sn: Surname
            - displayName: Full display name
            - password: User password
            - userPrincipalName: UPN with suffix
            - mail: Email address
            - telephoneNumber: Phone number
            - description: User description

    Note:
        This function expects Excel data with columns in this order:
        [First Name, Last Name, Additional Info, Password, Phone, Description]
        The function stops processing when it encounters a row with no last name.
        Russian names are transliterated to Latin characters for username generation.
    """
    user_data = []
    if not mail_domain:
        mail_domain = upn
    logger.debug(f"Mail_domain = {mail_domain}")
    for row in input_dict["data"]:
        if row[1] is not None:
            translit_name = str(
                transliterate.translit(row[1], language_code="ru", reversed=True)
            )
            cname = translit_name + str(row[2] or "")
            user = {
                "cn": cname,
                "givenName": row[0],
                "sn": row[1],
                "displayName": f"{row[0]} {row[1]}",
                "password": row[3],
                "userPrincipalName": f"{cname}@{upn}",
                "mail": f"{cname}@{mail_domain}",
                "telephoneNumber": row[4],
                "description": row[5],
            }
            user_data.append(user)
        else:
            break
    return user_data


def write_excel_data(path, data: list[dict]):
    """Write data to an Excel file with formatting.

    Creates a new Excel workbook and writes the provided data with consistent
    formatting including fonts and alignment.

    Args:
        path (str): Path where the Excel file will be saved.
        data (list[dict]): List of dictionaries to write. The first dictionary
            should contain all keys that will be used as column headers.

    Note:
        This function creates a new workbook and overwrites any existing file.
        Headers are written in the first row with Times New Roman 14pt bold font
        and single underline. Data rows use Times New Roman 14pt regular font.
        All cells are center-aligned. The function assumes all dictionaries
        in the data list have the same keys as the first dictionary.
    """
    wb = openpyxl.Workbook()
    sheet = wb.active
    header_font = Font(name="Times New Roman", size=14, bold=True, underline="single")
    body_font = Font(name="Times New Roman", size=14, bold=False)
    center_alignment = Alignment(horizontal="center")
    # headers
    for column in range(len(data[0])):
        sheet.cell(row=1, column=column + 1, value=list(data[0].keys())[column])
        sheet[1][column].font = header_font
        sheet[1][column].alignment = center_alignment

    # data
    for row in range(len(data)):
        for column in range(len(data[row])):
            sheet.cell(row + 2, column + 1, list(data[row].values())[column])
            sheet[row + 2][column].font = body_font
            sheet[row + 2][column].alignment = center_alignment
    wb.save(path)
